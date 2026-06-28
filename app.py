# ============================================================
# USMAN PUBLIC SCHOOL SYSTEM
# CAMPUS 45 - ANNUAL BUDGET FY 2026-27
# ============================================================

import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

st.set_page_config(
    page_title="Budget Dashboard 2026-27",
    layout="wide"
)

st.markdown("""
<style>

[data-testid="stSidebar"]{
    background-color:#1f4e78;
}

[data-testid="stSidebar"] *{
    color:white;
}

div.stDownloadButton > button{
    background-color:#1f4e78;
    color:white;
    border:none;
    border-radius:8px;
}

div.stDownloadButton > button:hover{
    background-color:#17375e;
    color:white;
}

</style>
""", unsafe_allow_html=True)

st.markdown(
    """
    <div style='text-align:center'>
        <h1>Usman Public School System</h1>
        <h2>Campus 45</h2>
        <h2 style='color:#1f4e78'>
            Annual Budget FY 2026-27
        </h2>
    </div>
    """,
    unsafe_allow_html=True
)

st.sidebar.header("Upload Files")

income_file = st.sidebar.file_uploader(
    "Upload Income File",
    type=["xlsx", "xls"]
)

expense_file = st.sidebar.file_uploader(
    "Upload Expense File",
    type=["xlsx", "xls"]
)


def read_excel_file(file, actual_sheet, budget_sheet):

    actual = pd.read_excel(
        file,
        sheet_name=actual_sheet
    )

    budget = pd.read_excel(
        file,
        sheet_name=budget_sheet
    )

    actual.columns = actual.columns.str.strip()
    budget.columns = budget.columns.str.strip()

    actual["Data Type"] = "Actual FY 2025-26"
    budget["Data Type"] = "Budget FY 2026-27"

    return actual, budget


def clean_data(df, category):

    df.columns = df.columns.str.strip()

    if "Amount" not in df.columns:
        numeric_cols = df.select_dtypes(
            include=np.number
        ).columns

        if len(numeric_cols) > 0:
            df = df.rename(
                columns={
                    numeric_cols[-1]: "Amount"
                }
            )

    if "Month" not in df.columns:
        df["Month"] = "Unknown"

    if "Particular" not in df.columns:
        df["Particular"] = df.iloc[:, 0]

    df["Category"] = category
    df["Amount"] = pd.to_numeric(
        df["Amount"],
        errors="coerce"
    ).fillna(0)

    return df


def convert_df_to_excel(df):

    output = BytesIO()

    with pd.ExcelWriter(
        output,
        engine="openpyxl"
    ) as writer:

        df.to_excel(
            writer,
            sheet_name="Profit & Loss",
            startrow=3,
            index=False
        )

        workbook = writer.book
        worksheet = writer.sheets["Profit & Loss"]

        worksheet.merge_cells("A1:C1")
        worksheet["A1"] = (
            "Usman Public School System\n"
            "Campus 45\n"
            "Annual Budget FY 2026-27"
        )

        worksheet["A1"].font = Font(
            size=16,
            bold=True
        )

        worksheet["A1"].alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

        blue_fill = PatternFill(
            start_color="1F4E78",
            end_color="1F4E78",
            fill_type="solid"
        )

        white_font = Font(
            color="FFFFFF",
            bold=True
        )

        thin = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        for cell in worksheet[4]:
            cell.fill = blue_fill
            cell.font = white_font
            cell.border = thin

        for row in worksheet.iter_rows(
            min_row=5,
            max_row=worksheet.max_row
        ):
            for cell in row:
                cell.border = thin

        worksheet.column_dimensions["A"].width = 25
        worksheet.column_dimensions["B"].width = 45
        worksheet.column_dimensions["C"].width = 20

    output.seek(0)
    return output
    if income_file and expense_file:

    try:

        income_actual, income_budget = read_excel_file(
            income_file,
            "Actual FY 2025-26",
            "Budget FY 2026-27"
        )

        expense_actual, expense_budget = read_excel_file(
            expense_file,
            "Actual FY 2025-26",
            "Budget FY 2026-27"
        )

        income_actual = clean_data(income_actual, "Income")
        income_budget = clean_data(income_budget, "Income")
        expense_actual = clean_data(expense_actual, "Expense")
        expense_budget = clean_data(expense_budget, "Expense")

        df = pd.concat(
            [
                income_actual,
                income_budget,
                expense_actual,
                expense_budget
            ],
            ignore_index=True
        )

        st.sidebar.header("Filters")

        months = [
            "Jul-25","Aug-25","Sep-25","Oct-25","Nov-25","Dec-25",
            "Jan-26","Feb-26","Mar-26","Apr-26","May-26","Jun-26",
            "Jul-26","Aug-26","Sep-26","Oct-26","Nov-26","Dec-26",
            "Jan-27","Feb-27","Mar-27","Apr-27","May-27","Jun-27"
        ]

        month_filter = st.sidebar.multiselect(
            "Select Month",
            months,
            default=months
        )

        category_filter = st.sidebar.multiselect(
            "Select Category",
            ["Income", "Expense"],
            default=["Income", "Expense"]
        )

        filtered_df = df[
            (df["Month"].astype(str).isin(month_filter))
            &
            (df["Category"].isin(category_filter))
        ]

        budget_df = filtered_df[
            filtered_df["Data Type"] == "Budget FY 2026-27"
        ]

        total_income = budget_df[
            budget_df["Category"] == "Income"
        ]["Amount"].sum()

        total_expense = budget_df[
            budget_df["Category"] == "Expense"
        ]["Amount"].sum()

        profit = total_income - total_expense

        profit_pct = (
            profit / total_income * 100
            if total_income != 0 else 0
        )

        c1, c2, c3, c4 = st.columns(4)

        c1.metric(
            "Total Income 2026-27",
            f"{total_income:,.0f}"
        )

        c2.metric(
            "Total Expense 2026-27",
            f"{total_expense:,.0f}"
        )

        c3.metric(
            "Profit / Loss",
            f"{profit:,.0f}"
        )

        c4.metric(
            "% Profit / Loss",
            f"{profit_pct:.2f}%"
        )

        st.divider()

        chart1, chart2 = st.columns(2)

        summary = (
            filtered_df
            .groupby(
                ["Data Type", "Category"]
            )["Amount"]
            .sum()
            .reset_index()
        )

        fig1 = px.bar(
            summary,
            x="Category",
            y="Amount",
            color="Data Type",
            barmode="group",
            title="Actual vs Budget"
        )

        chart1.plotly_chart(
            fig1,
            use_container_width=True
        )

        monthly_budget = (
            budget_df
            .groupby(
                ["Month", "Category"]
            )["Amount"]
            .sum()
            .reset_index()
        )

        fig2 = px.line(
            monthly_budget,
            x="Month",
            y="Amount",
            color="Category",
            markers=True,
            title="Month-wise Income vs Expense"
        )

        chart2.plotly_chart(
            fig2,
            use_container_width=True
        )

        pie_df = (
            budget_df
            .groupby("Category")["Amount"]
            .sum()
            .reset_index()
        )

        fig3 = px.pie(
            pie_df,
            names="Category",
            values="Amount",
            hole=0.5
        )

        st.plotly_chart(
            fig3,
            use_container_width=True
        )

        st.subheader("Top Budget Heads")

        col1, col2 = st.columns(2)

        income_top = (
            budget_df[budget_df["Category"] == "Income"]
            .groupby("Particular")["Amount"]
            .sum()
            .reset_index()
            .sort_values("Amount", ascending=False)
            .head(4)
        )

        expense_top = (
            budget_df[budget_df["Category"] == "Expense"]
            .groupby("Particular")["Amount"]
            .sum()
            .reset_index()
            .sort_values("Amount", ascending=False)
            .head(4)
        )

        with col1:
            st.subheader("Top 4 Income Heads")
            st.dataframe(
                income_top,
                use_container_width=True,
                hide_index=True
            )

        with col2:
            st.subheader("Top 4 Expense Heads")
            st.dataframe(
                expense_top,
                use_container_width=True,
                hide_index=True
            )

        pnl = (
            budget_df
            .groupby(
                ["Category", "Particular"]
            )["Amount"]
            .sum()
            .reset_index()
        )

        excel_file = convert_df_to_excel(pnl)

        st.download_button(
            "Download Profit & Loss Statement",
            data=excel_file,
            file_name="Profit_Loss_2026_27.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        st.error(f"Error: {e}")

else:
    st.info(
        "Please upload both Income and Expense files."
    )


